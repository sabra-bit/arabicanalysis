import streamlit as st
import openpyxl
import pandas as pd
import re
from wordcloud import WordCloud, STOPWORDS
import matplotlib.pyplot as plt
import plotly.express as px



st.title('arabic text analysis')
workbook = openpyxl.load_workbook(filename="datawithNoComment.xlsx", data_only=True)
sheet = workbook.active
data = []
for row in sheet.iter_rows():
  row_data = []
  for cell in row:
    row_data.append(cell.value)
  data.append(row_data)
data = pd.DataFrame(data)
data=data.iloc[1:]
data = data[[4,2]]
new_column_names = {4: 'text', 2: 'time'}
data = data.rename(columns=new_column_names)
data = data.dropna()
st.subheader("scraped data")
data

workbook1 = openpyxl.load_workbook(filename="drHalaData.xlsx", data_only=True)
sheet1 = workbook1.active
data1 = []
for row in sheet1.iter_rows():
  row_data = []
  for cell in row:
    row_data.append(cell.value)
  data1.append(row_data)

data1 = pd.DataFrame(data1)
data1=data1.iloc[1:]
data1 = data1[[1,2]]

new_column_names = {1: 'text', 2: 'time'}
data1 = data1.rename(columns=new_column_names)
data1 = data1.dropna()
st.subheader("old data")
data1



def clean_text(text, replace_data):
  text = re.sub(r"[^\u0600-\u06FF]", " ", str(text))
  text = re.sub(r"[^\w\s]", "", str(text))
  text = re.sub(r"(.)\1+", r"\1", str(text))

  # Apply word replacements
  for original_word, replacement in replace_data.items():
    text = text.replace(original_word, replacement)
  return text.strip()

def fetch_replace_data(excel_file, sheet_name):
  # Open the Excel workbook
  workbook = openpyxl.load_workbook(excel_file , data_only=True)
  sheet = workbook[sheet_name]

  # Read data from the sheet (assuming words are in first two columns)
  replace_data = {}
  for row in sheet.iter_rows(min_row=2):  # Skip the eheader row (assuming row 1)
    original_word = row[0].value
    replacement = row[1].value
    if original_word and replacement:  # Check for empty values
      replace_data[original_word] = replacement

  return replace_data

# Example usage
excel_file = 'replacements.xlsx'  # Replace with your actual file path
sheet_name = "Sheet1"  # Replace with your sheet name (if different)
replace_data = fetch_replace_data(excel_file, sheet_name)

cleaned_text_column = data['text'].apply(clean_text, args=(replace_data,))
data['Cleaned Text'] = cleaned_text_column

cleaned_text_column = data1['text'].apply(clean_text, args=(replace_data,))
data1['Cleaned Text'] = cleaned_text_column


file_path = "replace.xlsx"
df = pd.read_excel(file_path, sheet_name="Sheet2")
Enviromental  = df['Enviromental dimesnsion'].dropna()
Economic  = df['Economic Dimension'].dropna()
Social = df['Social dimension'].dropna()
rules = {
    "Enviromental": Enviromental.tolist(),
    "Economic": Economic.tolist(),
    "Social": Social.tolist()
}

def classify_text(text, rules):
    for class_name, keywords in rules.items():
        if any(keyword in text for keyword in keywords):
            return class_name
    return "unclassified"  # Default class for non-matching text


def classify_texttt(text, rules):
  for class_name, keywords in rules.items():
    for keyword in keywords:
      if keyword.lower() in text.lower():  
        return class_name, keyword
  return "unclassified", None 


result_df = pd.concat([data, data1], axis=0, ignore_index=True)
st.subheader("concatenate data")
result_df

result_df["class"] = result_df["Cleaned Text"].apply(lambda text: classify_texttt(text, rules)[0])
result_df["keyword"] = result_df["Cleaned Text"].apply(lambda text: classify_texttt(text, rules)[1])

options = st.multiselect(
    "choose the class",
    ["Enviromental", "Economic", "Social", "unclassified"],
    
)

if options:
    result_df = result_df[result_df['class'].isin(options)]
    result_df


    replications = result_df['class'].value_counts()
    replications

    def create_donut_chart(replications):
        # Create a donut chart using matplotlib
        plt.figure(figsize=(6, 6))
        plt.pie(replications, labels=replications.index, autopct='%1.1f%%', startangle=140, colors=plt.cm.Pastel1.colors, wedgeprops={'linewidth': 3, 'edgecolor': 'white'})
        plt.axis('equal')  # Equal aspect ratio for a circular pie plot
        plt.title('class Distribution')

        # Display the chart in Streamlit
        st.pyplot()
    create_donut_chart(replications)
    # fig = px.pie(replications, values='count', names='sentiment', hole=0.3)
    # fig.update_layout(title_text='sentiment')
    # st.plotly_chart(fig)



    def create_translation_dict(file_path, sheet_name):
        df = pd.read_excel(file_path, sheet_name=sheet_name)
        df = df[['word', 'translate']]
        return df.set_index('word')['translate'].to_dict()

    # Load the translation data into a dictionary
    translation_dict = create_translation_dict("translate.xlsx", "Sheet1")

    def translate_text(text):
        return translation_dict.get(text, text)  # Default to original text if not found

    # Apply the translation function to the 'keyword' column
    result_df['translation'] = result_df['keyword'].dropna().apply(translate_text)




    result_df['translation'] = result_df['translation'].str.lower()
    # Prepare text data
    text = " ".join(result_df["translation"].astype(str))

    from collections import Counter
    import nltk
    nltk.download('stopwords') 
    from nltk.corpus import stopwords

    # ... your code
    nltk.download('punkt')
    # Tokenize text
    words = nltk.word_tokenize(text)

    # Remove stop words
    stop_words = set(stopwords.words('english'))
    filtered_words = [word for word in words if word not in stop_words]

    # Create a new text string
    text = ' '.join(filtered_words)

    word_counts = Counter(text.split())

    # Generate word cloud
    wordcloud = WordCloud(width=800, height=400, background_color='white', stopwords=STOPWORDS, min_font_size=1).generate_from_frequencies(word_counts)
    # Display the word cloud
    st.subheader("trending word")
    st.set_option('deprecation.showPyplotGlobalUse', False)
    plt.figure(figsize=(10, 5))
    plt.imshow(wordcloud, interpolation='bilinear')
    plt.axis('off')
    plt.show()
    st.pyplot()
